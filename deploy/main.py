from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
import os

#custom font
free_type_compatible_font = os.path.abspath(
    os.path.join(
        os.path.dirname(__file__), 'D:/01. Kuliah/02. Angkatan/Pameran Konsep/deploy/font/Montserrat-Regular.ttf'
    )
)

# Load the Excel file
workbook = load_workbook('D:/01. Kuliah/02. Angkatan/Pameran Konsep/deploy/ready_push.xlsx')
sheet = workbook['Sheet1']  # Replace 'Sheet1' with the actual sheet name

# Load the certificate template image
template_image = Image.open('D:/01. Kuliah/02. Angkatan/Pameran Konsep/deploy/template_caption.png')  # Replace with your actual template file

# Specify the font and size for participant names
font    = ImageFont.truetype(font=free_type_compatible_font, size=30)  # Replace with your desired font and size

# Create desired folder
output_folder = 'D:/01. Kuliah/02. Angkatan/Pameran Konsep/deploy/archive'
os.makedirs(output_folder, exist_ok=True)

# Iterate through the rows of the Excel file
for row in sheet.iter_rows(min_row=2):  # Assuming the first row contains headers

    participant_n       = str(row[0].value)
    participant_name    = str.title(row[1].value)  # Assuming participant names are in the second column
    participant_title   = row[2].value
    participant_title   = participant_title.upper()
    participant_year    = str(row[3].value)
    participant_size    = str(row[4].value)
    participant_media   = row[5].value

    # Create a copy of the template image
    certificate = template_image.copy()

    # Add participant name to the certificate
    draw = ImageDraw.Draw(certificate)
    draw.text((90, 100), participant_name, fill='black', font=font)  # Adjust the coordinates as per your template
    draw.text((90, 200), participant_title, fill='black', font=font, stroke_width=1)
    draw.text((90, 250), participant_media, fill='black', font=font)
    draw.text((90, 300), participant_size, fill='black', font=font)
    draw.text((90, 350), participant_year, fill='black', font=font)


    # Save the certificate with a unique name
    certificate.save(os.path.join(output_folder, f'{participant_name}_caption_1.png'))  # Adjust the naming scheme as per your requirement

    #update the process
    print(f'{participant_name}_{participant_n} is done')

    #certificate 2

    certificate_2 = template_image.copy()
    draw = ImageDraw.Draw(certificate_2)

    participant_title   = row[6].value
    participant_title   = participant_title.upper()
    participant_year    = str(row[7].value)
    participant_size    = str(row[8].value)
    participant_media   = row[9].value

    draw.text((90, 100), participant_name, fill='black', font=font)  # Adjust the coordinates as per your template
    draw.text((90, 200), participant_title, fill='black', font=font, stroke_width=1)
    draw.text((90, 250), participant_media, fill='black', font=font)
    draw.text((90, 300), participant_size, fill='black', font=font)
    draw.text((90, 350), participant_year, fill='black', font=font)

    certificate_2.save(os.path.join(output_folder, f'{participant_name}_caption_2.png'))

    print(f'{participant_name} 2 is done')

    #certificate 3

    certificate_3 = template_image.copy()
    draw = ImageDraw.Draw(certificate_3)

    participant_title   = row[10].value
    participant_title   = participant_title.upper()
    participant_year    = str(row[11].value)
    participant_size    = str(row[12].value)
    participant_media   = row[13].value

    draw.text((90, 100), participant_name, fill='black', font=font)  # Adjust the coordinates as per your template
    draw.text((90, 200), participant_title, fill='black', font=font, stroke_width=1)
    draw.text((90, 250), participant_media, fill='black', font=font)
    draw.text((90, 300), participant_size, fill='black', font=font)
    draw.text((90, 350), participant_year, fill='black', font=font)

    certificate_3.save(os.path.join(output_folder, f'{participant_name}_caption_3.png'))

    print(f'{participant_name} 3 is done')

    #certificate 4

    certificate_4 = template_image.copy()
    draw = ImageDraw.Draw(certificate_4)

    participant_title   = row[14].value
    participant_title   = participant_title.upper()
    participant_year    = str(row[15].value)
    participant_size    = str(row[16].value)
    participant_media   = row[17].value

    draw.text((90, 100), participant_name, fill='black', font=font)  # Adjust the coordinates as per your template
    draw.text((90, 200), participant_title, fill='black', font=font, stroke_width=1)
    draw.text((90, 250), participant_media, fill='black', font=font)
    draw.text((90, 300), participant_size, fill='black', font=font)
    draw.text((90, 350), participant_year, fill='black', font=font)

    certificate_4.save(os.path.join(output_folder, f'{participant_name}_caption_4.png'))

    print(f'{participant_name} 4 is done')


